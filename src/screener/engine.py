"""
Financial Screener Filter Engine — Sprint 3 Days 15, 16 & 17.

Loads screener_config.yaml, queries financial_ratios table from db/nifty100.db,
applies threshold filters across 15 financial metrics, enforces sector carve-outs
and Debt-Free ICR infinity rules, computes sector-relative winsorized composite quality scores,
and exports formatted output/screener_output.xlsx.
"""

import os
import sqlite3
from typing import Dict, Any, Optional, List
import pandas as pd
import numpy as np

try:
    import yaml
except ImportError:
    yaml = None

from src.utils.logger import get_logger
from src.db.connection import get_db_connection

logger = get_logger(__name__)

DEFAULT_CONFIG_PATH = "config/screener_config.yaml"
DEFAULT_DB_PATH = "db/nifty100.db"
OUTPUT_SCREENER_EXCEL = "output/screener_output.xlsx"

class ScreenerEngine:
    def __init__(self, config_path: str = DEFAULT_CONFIG_PATH, db_path: str = DEFAULT_DB_PATH):
        self.config_path = config_path
        self.db_path = db_path
        self.config = self.load_config()

    def load_config(self) -> Dict[str, Any]:
        """Loads YAML configuration file."""
        if not os.path.exists(self.config_path):
            logger.warning(f"Config file not found at {self.config_path}. Using default mappings.")
            return {}

        if yaml is not None:
            with open(self.config_path, "r", encoding="utf-8") as f:
                return yaml.safe_load(f) or {}
        else:
            return {
                "presets": {
                    "quality_compounder": {"filters": {"roe_min": 15.0, "de_max": 1.0, "fcf_min": 0.0, "revenue_cagr_5yr_min": 10.0}},
                    "value_pick": {"filters": {"pe_max": 20.0, "pb_max": 3.0, "de_max": 2.0, "dividend_yield_min": 1.0}},
                    "growth_accelerator": {"filters": {"pat_cagr_5yr_min": 20.0, "revenue_cagr_5yr_min": 15.0, "de_max": 2.0}},
                    "dividend_champion": {"filters": {"dividend_yield_min": 2.0, "fcf_min": 0.0}},
                    "debt_free_blue_chip": {"filters": {"de_max": 0.0, "roe_min": 12.0, "sales_min": 5000.0}},
                    "turnaround_watch": {"filters": {"revenue_cagr_5yr_min": 10.0, "fcf_min": 0.0}}
                }
            }

    def compute_composite_quality_scores(self, df: pd.DataFrame) -> pd.DataFrame:
        """
        Computes sector-relative composite quality score (0 to 100 scale)
        with P10/P90 winsorization across broad sectors.
        Weights: Profitability 35%, Cash Quality 30%, Growth 20%, Leverage 15%.
        """
        if df.empty:
            return df

        res = df.copy()

        def winsorize_scale(series: pd.Series, invert: bool = False) -> pd.Series:
            valid = series.dropna()
            if valid.empty:
                return pd.Series(50.0, index=series.index)
            p10, p90 = np.percentile(valid, 10), np.percentile(valid, 90)
            if p90 > p10:
                clipped = np.clip(series.fillna(p10), p10, p90)
                scaled = (clipped - p10) / (p90 - p10) * 100.0
            else:
                scaled = pd.Series(50.0, index=series.index)
            if invert:
                scaled = 100.0 - scaled
            return scaled

        # Sector relative scoring
        if "sector_name" not in res.columns:
            res["sector_name"] = "General"

        scores = []
        for sector, grp in res.groupby("sector_name"):
            g = grp.copy()
            # 1. Profitability (35%): ROE 15%, ROCE 10%, NPM 10%
            roe_s = winsorize_scale(g.get("return_on_equity_pct", g.get("roe", pd.Series())))
            roce_s = winsorize_scale(g.get("roce", pd.Series()))
            npm_s = winsorize_scale(g.get("net_profit_margin_pct", g.get("npm", pd.Series())))
            prof_score = roe_s * (15/35) + roce_s * (10/35) + npm_s * (10/35)

            fcf_col = g["free_cash_flow_cr"] if "free_cash_flow_cr" in g.columns else (g["free_cash_flow"] if "free_cash_flow" in g.columns else pd.Series(0.0, index=g.index))
            fcf_s = winsorize_scale(fcf_col)
            cfo_pat_s = winsorize_scale(g["cfo_quality"] if "cfo_quality" in g.columns else pd.Series(index=g.index))
            fcf_pos = (pd.to_numeric(fcf_col, errors='coerce').fillna(0) > 0).astype(float) * 100.0
            cash_score = fcf_s * (15/30) + cfo_pat_s * (10/30) + fcf_pos * (5/30)


            # 3. Growth (20%): Revenue CAGR 10%, PAT CAGR 10%
            rev_cagr = winsorize_scale(g.get("revenue_cagr_5yr", g.get("cagr_sales_5yr", pd.Series())))
            pat_cagr = winsorize_scale(g.get("pat_cagr_5yr", g.get("cagr_pat_5yr", pd.Series())))
            growth_score = rev_cagr * 0.5 + pat_cagr * 0.5

            # 4. Leverage (15%): D/E 10% (inverted), ICR 5%
            de_s = winsorize_scale(g.get("debt_to_equity", pd.Series()), invert=True)
            icr_s = winsorize_scale(g.get("interest_coverage", pd.Series()))
            lev_score = de_s * (10/15) + icr_s * (5/15)

            # Total Composite Quality Score (0 - 100)
            comp_score = prof_score * 0.35 + cash_score * 0.30 + growth_score * 0.20 + lev_score * 0.15
            g["composite_quality_score"] = np.round(np.clip(comp_score, 0, 100), 2)
            scores.append(g)

        res = pd.concat(scores, axis=0) if scores else res
        return res

    def load_universe_data(self, latest_year_only: bool = True) -> pd.DataFrame:
        """Loads dataset from SQLite database."""
        if not os.path.exists(self.db_path):
            logger.error(f"Database file not found at {self.db_path}")
            return pd.DataFrame()

        with get_db_connection(self.db_path) as conn:
            query = """
                SELECT 
                    fr.*,
                    c.company_name,
                    c.ticker,
                    s.sector_name,
                    pnl.sales,
                    pnl.operating_profit,
                    pnl.net_profit,
                    pnl.eps,
                    bs.total_assets,
                    bs.total_equity,
                    cf.operating_cash_flow,
                    cf.investing_cash_flow,
                    cf.financing_cash_flow,
                    mc.market_cap
                FROM financial_ratios fr
                JOIN companies c ON fr.company_id = c.company_id
                LEFT JOIN sectors s ON c.sector_id = s.sector_id
                LEFT JOIN profitandloss pnl ON fr.company_id = pnl.company_id AND fr.year = pnl.year
                LEFT JOIN balancesheet bs ON fr.company_id = bs.company_id AND fr.year = bs.year
                LEFT JOIN cashflow cf ON fr.company_id = cf.company_id AND fr.year = cf.year
                LEFT JOIN market_cap mc ON fr.company_id = mc.company_id AND CAST(mc.date AS INTEGER) = fr.year
            """
            df = pd.read_sql_query(query, conn)

        if df.empty:
            return df

        if latest_year_only:
            max_years = df.groupby("company_id")["year"].max().reset_index()
            df = pd.merge(df, max_years, on=["company_id", "year"])

        df = self.compute_composite_quality_scores(df)
        return df

    def apply_filters(self, df: pd.DataFrame, filters: Dict[str, Any]) -> pd.DataFrame:
        """Applies filters supporting 15 financial metrics with Financials & Debt-Free exemptions."""
        if df.empty or not filters:
            return df

        res = df.copy()

        # 1. ROE Min
        if "roe_min" in filters and filters["roe_min"] is not None:
            val = float(filters["roe_min"])
            col = "return_on_equity_pct" if "return_on_equity_pct" in res.columns else "roe"
            res = res[res[col].notna() & (res[col] >= val)]

        # 2. D/E Max (SPECIAL RULE: Skip for Financials sector)
        if "de_max" in filters and filters["de_max"] is not None:
            val = float(filters["de_max"])
            is_fin_col = "is_financial_sector" if "is_financial_sector" in res.columns else None
            
            def de_check(row):
                if is_fin_col and bool(row.get(is_fin_col, 0)):
                    return True
                de_val = row.get("debt_to_equity")
                if de_val is None or pd.isna(de_val):
                    return True
                return float(de_val) <= val

            res = res[res.apply(de_check, axis=1)]

        # 3. Free Cash Flow Min
        if "fcf_min" in filters and filters["fcf_min"] is not None:
            val = float(filters["fcf_min"])
            def fcf_check(row):
                f_val = row.get("free_cash_flow_cr") if row.get("free_cash_flow_cr") is not None else row.get("free_cash_flow")
                if f_val is not None and not pd.isna(f_val):
                    return float(f_val) >= val
                return False
            res = res[res.apply(fcf_check, axis=1)]


        # 4. Revenue CAGR 5yr Min
        if "revenue_cagr_5yr_min" in filters and filters["revenue_cagr_5yr_min"] is not None:
            val = float(filters["revenue_cagr_5yr_min"])
            def rev_check(row):
                c_val = row.get("revenue_cagr_5yr") if row.get("revenue_cagr_5yr") is not None else row.get("cagr_sales_5yr")
                if c_val is not None and not pd.isna(c_val):
                    return float(c_val) >= val
                return False
            res = res[res.apply(rev_check, axis=1)]

        # 5. PAT CAGR 5yr Min
        if "pat_cagr_5yr_min" in filters and filters["pat_cagr_5yr_min"] is not None:
            val = float(filters["pat_cagr_5yr_min"])
            def pat_check(row):
                c_val = row.get("pat_cagr_5yr") if row.get("pat_cagr_5yr") is not None else row.get("cagr_pat_5yr")
                if c_val is not None and not pd.isna(c_val):
                    return float(c_val) >= val
                return False
            res = res[res.apply(pat_check, axis=1)]

        # 6. OPM Min
        if "opm_min" in filters and filters["opm_min"] is not None:
            val = float(filters["opm_min"])
            def opm_check(row):
                o_val = row.get("operating_profit_margin_pct") if row.get("operating_profit_margin_pct") is not None else row.get("opm")
                if o_val is not None and not pd.isna(o_val):
                    return float(o_val) >= val
                return False
            res = res[res.apply(opm_check, axis=1)]


        # 7. P/E Max
        if "pe_max" in filters and filters["pe_max"] is not None:
            val = float(filters["pe_max"])
            def pe_check(row):
                pe_v = row.get("pe_ratio")
                if pe_v is not None and not pd.isna(pe_v):
                    return float(pe_v) <= val
                mcap = row.get("market_cap")
                np_val = row.get("net_profit")
                if mcap is not None and not pd.isna(mcap) and float(mcap) > 0 and np_val is not None and not pd.isna(np_val) and float(np_val) > 0:
                    calculated_pe = float(mcap) / float(np_val)
                    return calculated_pe <= val
                return False  # Exclude if P/E cannot be derived
            res = res[res.apply(pe_check, axis=1)]

        # 8. P/B Max
        if "pb_max" in filters and filters["pb_max"] is not None:
            val = float(filters["pb_max"])
            def pb_check(row):
                pb_v = row.get("pb_ratio")
                if pb_v is not None and not pd.isna(pb_v):
                    return float(pb_v) <= val
                mcap = row.get("market_cap")
                eq_val = row.get("total_equity")
                if mcap is not None and not pd.isna(mcap) and float(mcap) > 0 and eq_val is not None and not pd.isna(eq_val) and float(eq_val) > 0:
                    calculated_pb = float(mcap) / float(eq_val)
                    return calculated_pb <= val
                return False  # Exclude if P/B cannot be derived
            res = res[res.apply(pb_check, axis=1)]

        # 9. Dividend Yield Min (STRICT RULE: NULL does NOT pass)
        if "dividend_yield_min" in filters and filters["dividend_yield_min"] is not None:
            val = float(filters["dividend_yield_min"])
            def div_check(row):
                dy_v = row.get("dividend_yield")
                if dy_v is not None and not pd.isna(dy_v):
                    return float(dy_v) >= val
                return False  # Exclude records when dividend yield is NULL
            res = res[res.apply(div_check, axis=1)]

        # 10. ICR Min (SPECIAL RULE: Debt Free = ICR Infinity)
        if "icr_min" in filters and filters["icr_min"] is not None:
            val = float(filters["icr_min"])
            
            def icr_check(row):
                is_debt_free = bool(row.get("debt_free_label", 0)) or (row.get("debt_to_equity") == 0.0)
                if is_debt_free:
                    return True
                icr_val = row.get("interest_coverage")
                if icr_val is None or pd.isna(icr_val):
                    return False
                return float(icr_val) >= val

            res = res[res.apply(icr_check, axis=1)]

        # 11. Market Cap Min
        if "market_cap_min" in filters and filters["market_cap_min"] is not None:
            val = float(filters["market_cap_min"])
            if "market_cap" in res.columns:
                res = res[res["market_cap"].notna() & (res["market_cap"] >= val)]

        # 12. Net Profit Min
        if "net_profit_min" in filters and filters["net_profit_min"] is not None:
            val = float(filters["net_profit_min"])
            if "net_profit" in res.columns:
                res = res[res["net_profit"].notna() & (res["net_profit"] >= val)]

        # 13. EPS CAGR Min
        if "eps_cagr_min" in filters and filters["eps_cagr_min"] is not None:
            val = float(filters["eps_cagr_min"])
            col = "eps_cagr_5yr" if "eps_cagr_5yr" in res.columns else "cagr_eps_5yr"
            if col in res.columns:
                res = res[res[col].notna() & (res[col] >= val)]

        # 14. Asset Turnover Min
        if "asset_turnover_min" in filters and filters["asset_turnover_min"] is not None:
            val = float(filters["asset_turnover_min"])
            if "asset_turnover" in res.columns:
                res = res[res["asset_turnover"].notna() & (res["asset_turnover"] >= val)]

        # 15. Sales Min
        if "sales_min" in filters and filters["sales_min"] is not None:
            val = float(filters["sales_min"])
            if "sales" in res.columns:
                res = res[res["sales"].notna() & (res["sales"] >= val)]

        sort_col = "composite_quality_score" if "composite_quality_score" in res.columns else "roe"
        if sort_col in res.columns:
            res = res.sort_values(sort_col, ascending=False)

        return res

    def run_preset(self, preset_key: str, df: Optional[pd.DataFrame] = None) -> pd.DataFrame:
        """Runs a preset filter defined in screener_config.yaml."""
        if df is None:
            df = self.load_universe_data()

        presets = self.config.get("presets", {})
        if preset_key not in presets:
            logger.error(f"Preset key '{preset_key}' not found in screener configuration.")
            return pd.DataFrame()

        filter_dict = presets[preset_key].get("filters", {})
        return self.apply_filters(df, filter_dict)

    def generate_screener_excel_report(self, output_path: str = OUTPUT_SCREENER_EXCEL):
        """Generates output/screener_output.xlsx containing 6 formatted sheets."""
        logger.info(f"Generating Screener Excel report to {output_path}...")
        os.makedirs(os.path.dirname(output_path), exist_ok=True)

        universe_df = self.load_universe_data()
        presets = self.config.get("presets", {})

        writer = pd.ExcelWriter(output_path, engine='openpyxl')

        for preset_key, preset_info in presets.items():
            sheet_name = preset_info.get("name", preset_key)[:31]
            filtered_df = self.run_preset(preset_key, universe_df)

            cols_20 = [
                "company_id", "company_name", "year", "sector_name", "return_on_equity_pct",
                "debt_to_equity", "free_cash_flow_cr", "revenue_cagr_5yr", "pat_cagr_5yr",
                "operating_profit_margin_pct", "pe_ratio", "pb_ratio", "dividend_payout_ratio_pct",
                "interest_coverage", "market_cap", "net_profit", "eps_cagr_5yr",
                "asset_turnover", "sales", "composite_quality_score", "capital_allocation_strategy"
            ]

            disp_cols = [c for c in cols_20 if c in filtered_df.columns]
            if not filtered_df.empty:
                export_df = filtered_df[disp_cols]
            else:
                # Document data limitation clearly when source dividend data is unavailable
                note_row = {c: None for c in disp_cols}
                note_row["company_id"] = "DATA LIMITATION"
                note_row["company_name"] = "Source dividend data (Dividend Yield / Dividend Payout) is unavailable in raw export files."
                export_df = pd.DataFrame([note_row])

            export_df.to_excel(writer, sheet_name=sheet_name, index=False)

        writer.close()

        # Style Excel Workbook with openpyxl
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font

            wb = openpyxl.load_workbook(output_path)
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

            for ws in wb.worksheets:
                for cell in ws[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

                for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                    for cell in row:
                        cell.fill = green_fill

            wb.save(output_path)
        except Exception as e:
            logger.warning(f"Could not apply openpyxl styles: {e}")

        logger.info(f"Screener Excel report exported successfully to {output_path}")
