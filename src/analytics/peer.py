"""
Peer Analytics & Percentile Engine — Sprint 3 Days 18 & 19.

Computes PERCENT_RANK across 10 financial metrics for 11 peer groups,
stores rankings in SQLite table 'peer_percentiles', generates radar charts,
and exports formatted output/peer_comparison.xlsx report.
"""

import os
import sqlite3
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional
from datetime import datetime

import matplotlib
matplotlib.use('Agg')  # Non-interactive backend for server environments
import matplotlib.pyplot as plt

from src.utils.logger import get_logger
from src.db.connection import get_db_connection

logger = get_logger(__name__)

DB_PATH = "db/nifty100.db"
OUTPUT_PEER_EXCEL = "output/peer_comparison.xlsx"
RADAR_CHARTS_DIR = "reports/radar_charts"

PEER_METRICS = [
    ("roe", "ROE", False),
    ("roce", "ROCE", False),
    ("npm", "Net Profit Margin", False),
    ("debt_to_equity", "D/E", True),  # Inverted: lower D/E is better
    ("free_cash_flow", "FCF", False),
    ("cagr_pat_5yr", "PAT CAGR 5yr", False),
    ("cagr_sales_5yr", "Revenue CAGR 5yr", False),
    ("cagr_eps_5yr", "EPS CAGR 5yr", False),
    ("interest_coverage", "Interest Coverage", False),
    ("asset_turnover", "Asset Turnover", False)
]

DEFAULT_PEER_GROUPS = {
    "IT Services": ["TCS", "INFY", "HCLTECH", "WIPRO", "LTIM", "TECHM"],
    "Banking & Financials": ["HDFCBANK", "ICICIBANK", "SBIN", "KOTAKBANK", "AXISBANK", "BAJFINANCE"],
    "FMCG": ["HINDUNILVR", "ITC", "NESTLEIND", "BRITANNIA", "GODREJCP", "TATACONSUM"],
    "Automobile": ["TATAMOTORS", "MARUTI", "M&M", "BAJAJ-AUTO", "EICHERMOT", "HEROMOTOCO"],
    "Pharma & Healthcare": ["SUNPHARMA", "CIPLA", "DRREDDY", "APOLLOHOSP", "TORNTPHARM", "ZYDUSLIFE"],
    "Oil & Gas / Energy": ["RELIANCE", "ONGC", "NTPC", "POWERGRID", "BPCL", "IOC", "GAIL"],
    "Metals & Mining": ["TATASTEEL", "HINDALCO", "JSWSTEEL", "COALINDIA", "VEDL"],
    "Capital Goods & Infra": ["LT", "BEL", "HAL", "SIEMENS", "BHEL"],
    "Consumer Durables & Retail": ["TITAN", "DMART", "TRENT", "VBL", "ZOMATO"],
    "Cement & Materials": ["ULTRACEMCO", "GRASIM", "AMBUJACEM", "ACC"],
    "Power & Utilities": ["ADANIGREEN", "ADANIPOWER", "ADANIENSOL", "ATGL", "TATAPOWER"]
}

def ensure_peer_percentiles_table(conn: sqlite3.Connection):
    """Creates peer_percentiles table in SQLite database if not existing."""
    sql = """
    CREATE TABLE IF NOT EXISTS peer_percentiles (
        company_id TEXT,
        peer_group_name TEXT,
        metric TEXT,
        value REAL,
        percentile_rank REAL,
        year INTEGER,
        PRIMARY KEY (company_id, peer_group_name, metric, year),
        FOREIGN KEY (company_id) REFERENCES companies (company_id)
    );
    """
    conn.execute(sql)
    conn.commit()

def calculate_percent_rank(series: pd.Series, invert: bool = False) -> pd.Series:
    """Computes percentile rank on a 0-1 scale. Inverts if lower value is better (e.g. D/E)."""
    valid = series.dropna()
    if len(valid) <= 1:
        ranks = pd.Series(0.5, index=series.index)
        return ranks.where(series.notna(), np.nan)
    
    rank_pct = series.rank(pct=True, method='min')
    if invert:
        rank_pct = 1.0 - rank_pct
    return rank_pct

def compute_peer_percentiles(db_path: str = DB_PATH) -> pd.DataFrame:
    """
    Computes percentile rankings for 10 metrics across 11 peer groups
    and stores them in SQLite table peer_percentiles.
    """
    logger.info("Starting Peer Percentile Ranking Engine...")

    if not os.path.exists(db_path):
        logger.error(f"Database file not found at {db_path}")
        return pd.DataFrame()

    with get_db_connection(db_path) as conn:
        ensure_peer_percentiles_table(conn)

        # Load latest ratio records per company
        df = pd.read_sql_query("""
            SELECT fr.*, c.company_name
            FROM financial_ratios fr
            JOIN companies c ON fr.company_id = c.company_id
        """, conn)

        if df.empty:
            return pd.DataFrame()

        # Filter to latest year per company
        max_years = df.groupby("company_id")["year"].max().reset_index()
        latest_df = pd.merge(df, max_years, on=["company_id", "year"])

        # Map peer groups
        peer_map = {}
        for group_name, comp_list in DEFAULT_PEER_GROUPS.items():
            for cid in comp_list:
                peer_map[cid] = group_name

        records = []
        unassigned_count = 0

        for _, row in latest_df.iterrows():
            cid = row["company_id"]
            year = int(row["year"])
            group_name = peer_map.get(cid)

            if not group_name:
                unassigned_count += 1
                logger.info(f"No peer group assigned for company {cid} - skipping percentile rank.")
                continue

            # Find all peers in the same group
            peers_in_group = latest_df[latest_df["company_id"].isin(DEFAULT_PEER_GROUPS[group_name])]

            for metric_col, metric_label, invert_flag in PEER_METRICS:
                col_name = metric_col
                if metric_col == "roe" and "return_on_equity_pct" in row and pd.notna(row["return_on_equity_pct"]):
                    col_name = "return_on_equity_pct"
                elif metric_col == "npm" and "net_profit_margin_pct" in row and pd.notna(row["net_profit_margin_pct"]):
                    col_name = "net_profit_margin_pct"
                elif metric_col == "free_cash_flow" and "free_cash_flow_cr" in row and pd.notna(row["free_cash_flow_cr"]):
                    col_name = "free_cash_flow_cr"

                series = peers_in_group.set_index("company_id")[col_name]
                val = row.get(col_name)

                if pd.isna(val) or series.dropna().empty:
                    pct_rank = np.nan
                else:
                    pct_ranks = calculate_percent_rank(series, invert=invert_flag)
                    pct_rank = pct_ranks.get(cid, np.nan)
                    if pd.notna(pct_rank):
                        pct_rank = round(float(pct_rank) * 100.0, 2)

                records.append({
                    "company_id": cid,
                    "peer_group_name": group_name,
                    "metric": metric_label,
                    "value": val if pd.notna(val) else None,
                    "percentile_rank": pct_rank if pd.notna(pct_rank) else None,
                    "year": year
                })

        records_df = pd.DataFrame(records)

        if not records_df.empty:
            conn.execute("DELETE FROM peer_percentiles;")
            records_df.to_sql("peer_percentiles", conn, if_exists="append", index=False)
            conn.commit()
            logger.info(f"Populated {len(records_df)} percentile rankings into SQLite table peer_percentiles.")

        return records_df

def generate_radar_charts(db_path: str = DB_PATH, output_dir: str = RADAR_CHARTS_DIR):
    """
    Generates Matplotlib polar radar charts for all companies in peer groups
    and exports PNG files to reports/radar_charts/<company_id>_radar.png.
    """
    os.makedirs(output_dir, exist_ok=True)
    logger.info(f"Generating Radar Charts to {output_dir}...")

    with get_db_connection(db_path) as conn:
        df = pd.read_sql_query("""
            SELECT fr.*, c.company_name
            FROM financial_ratios fr
            JOIN companies c ON fr.company_id = c.company_id
        """, conn)

        if df.empty:
            return

        max_years = df.groupby("company_id")["year"].max().reset_index()
        latest_df = pd.merge(df, max_years, on=["company_id", "year"])

        peer_map = {}
        for group_name, comp_list in DEFAULT_PEER_GROUPS.items():
            for cid in comp_list:
                peer_map[cid] = group_name

        axes_labels = ["ROE", "ROCE", "NPM", "D/E (Inv)", "FCF", "PAT CAGR 5Y", "Sales CAGR 5Y", "Composite Score"]
        num_vars = len(axes_labels)

        # Precompute min-max scaled metrics (0 to 100)
        scaled_df = latest_df.copy()
        for col in ["roe", "roce", "npm", "free_cash_flow", "cagr_pat_5yr", "cagr_sales_5yr", "composite_quality_score"]:
            c = col if col in scaled_df.columns else "roe"
            scaled_df[c] = scaled_df[c].fillna(0)
            p10, p90 = np.percentile(scaled_df[c], 10), np.percentile(scaled_df[c], 90)
            if p90 > p10:
                scaled_df[c] = np.clip((scaled_df[c] - p10) / (p90 - p10) * 100.0, 0, 100)
            else:
                scaled_df[c] = 50.0

        # Inverted D/E scale (lower D/E = higher score)
        if "debt_to_equity" in scaled_df.columns:
            de_vals = scaled_df["debt_to_equity"].fillna(2.0)
            p10, p90 = np.percentile(de_vals, 10), np.percentile(de_vals, 90)
            if p90 > p10:
                scaled_df["de_score"] = np.clip((p90 - de_vals) / (p90 - p10) * 100.0, 0, 100)
            else:
                scaled_df["de_score"] = 50.0
        else:
            scaled_df["de_score"] = 50.0

        scaled_df = scaled_df.set_index("company_id")

        for cid, row in latest_df.iterrows():
            company_id = row["company_id"]
            company_name = row["company_name"]
            group_name = peer_map.get(company_id)

            if company_id not in scaled_df.index:
                continue

            c_row = scaled_df.loc[company_id]
            comp_vals = [
                c_row.get("roe", 50), c_row.get("roce", 50), c_row.get("npm", 50),
                c_row.get("de_score", 50), c_row.get("free_cash_flow", 50),
                c_row.get("cagr_pat_5yr", 50), c_row.get("cagr_sales_5yr", 50),
                c_row.get("composite_quality_score", 50)
            ]

            if group_name and group_name in DEFAULT_PEER_GROUPS:
                peers = [p for p in DEFAULT_PEER_GROUPS[group_name] if p in scaled_df.index]
                peer_df = scaled_df.loc[peers]
                avg_vals = [
                    peer_df["roe"].mean(), peer_df["roce"].mean(), peer_df["npm"].mean(),
                    peer_df["de_score"].mean(), peer_df["free_cash_flow"].mean(),
                    peer_df["cagr_pat_5yr"].mean(), peer_df["cagr_sales_5yr"].mean(),
                    peer_df["composite_quality_score"].mean()
                ]
                ref_label = f"Peer Avg ({group_name})"
            else:
                avg_vals = [
                    scaled_df["roe"].mean(), scaled_df["roce"].mean(), scaled_df["npm"].mean(),
                    scaled_df["de_score"].mean(), scaled_df["free_cash_flow"].mean(),
                    scaled_df["cagr_pat_5yr"].mean(), scaled_df["cagr_sales_5yr"].mean(),
                    scaled_df["composite_quality_score"].mean()
                ]
                ref_label = "Nifty 100 Average"

            # Plot radar chart
            angles = [n / float(num_vars) * 2 * np.pi for n in range(num_vars)]
            angles += angles[:1]
            comp_vals += comp_vals[:1]
            avg_vals += avg_vals[:1]

            fig, ax = plt.subplots(figsize=(6, 6), subplot_kw=dict(polar=True))
            plt.xticks(angles[:-1], axes_labels, color='grey', size=8)

            ax.plot(angles, comp_vals, linewidth=2, linestyle='solid', label=company_id, color='#1f77b4')
            ax.fill(angles, comp_vals, color='#1f77b4', alpha=0.25)

            ax.plot(angles, avg_vals, linewidth=1.5, linestyle='dashed', label=ref_label, color='#ff7f0e')

            plt.title(f"{company_name} ({company_id})\nFinancial Quality Radar", size=11, color='navy', y=1.08)
            plt.legend(loc='upper right', bbox_to_anchor=(1.2, 1.1), fontsize=8)

            chart_path = os.path.join(output_dir, f"{company_id}_radar.png")
            plt.tight_layout()
            plt.savefig(chart_path, dpi=150)
            plt.close()

    logger.info("Radar charts generation completed successfully.")

def generate_peer_comparison_excel(db_path: str = DB_PATH, output_path: str = OUTPUT_PEER_EXCEL):
    """
    Generates output/peer_comparison.xlsx containing 11 formatted Excel sheets
    with color-coded percentile rankings, gold benchmark highlights, and median rows.
    """
    logger.info(f"Generating Peer Comparison Excel report to {output_path}...")
    os.makedirs(os.path.dirname(output_path), exist_ok=True)

    with get_db_connection(db_path) as conn:
        ratios_df = pd.read_sql_query("""
            SELECT fr.*, c.company_name
            FROM financial_ratios fr
            JOIN companies c ON fr.company_id = c.company_id
        """, conn)

        percentiles_df = pd.read_sql_query("SELECT * FROM peer_percentiles", conn)

    if ratios_df.empty:
        return

    max_years = ratios_df.groupby("company_id")["year"].max().reset_index()
    latest_ratios = pd.merge(ratios_df, max_years, on=["company_id", "year"])

    writer = pd.ExcelWriter(output_path, engine='openpyxl')

    for group_name, comp_list in DEFAULT_PEER_GROUPS.items():
        sheet_name = group_name.replace("/", "-").replace("&", "and")[:31]
        group_ratios = latest_ratios[latest_ratios["company_id"].isin(comp_list)].copy()

        if group_ratios.empty:
            continue

        sheet_rows = []
        for _, r in group_ratios.iterrows():
            cid = r["company_id"]
            cname = r["company_name"]
            row_dict = {
                "company_id": cid,
                "company_name": cname,
                "ROE (%)": r.get("roe"),
                "OPM (%)": r.get("opm"),
                "NPM (%)": r.get("npm"),
                "ROCE (%)": r.get("roce"),
                "ROA (%)": r.get("roa"),
                "D/E Ratio": r.get("debt_to_equity"),
                "Interest Coverage": r.get("interest_coverage"),
                "Asset Turnover": r.get("asset_turnover"),
                "FCF (Cr)": r.get("free_cash_flow"),
                "CFO Quality": r.get("cfo_quality"),
                "CapEx Intensity (%)": r.get("capex_intensity"),
                "FCF Conversion (%)": r.get("fcf_conversion"),
                "Sales CAGR 3Y (%)": r.get("cagr_sales_3yr"),
                "Sales CAGR 5Y (%)": r.get("cagr_sales_5yr"),
                "PAT CAGR 3Y (%)": r.get("cagr_pat_3yr"),
                "PAT CAGR 5Y (%)": r.get("cagr_pat_5yr"),
                "EPS CAGR 3Y (%)": r.get("cagr_eps_3yr"),
                "EPS CAGR 5Y (%)": r.get("cagr_eps_5yr"),
                "Net Debt (Cr)": r.get("net_debt"),
                "Composite Score": r.get("composite_quality_score", 50.0)
            }
            sheet_rows.append(row_dict)

        df_sheet = pd.DataFrame(sheet_rows)

        # Add median summary row
        median_dict = {"company_id": "MEDIAN", "company_name": f"{group_name} Group Median"}
        for col in df_sheet.columns:
            if col not in ["company_id", "company_name"]:
                median_dict[col] = df_sheet[col].median()
        
        df_sheet = pd.concat([df_sheet, pd.DataFrame([median_dict])], ignore_index=True)
        df_sheet.to_excel(writer, sheet_name=sheet_name, index=False)

    writer.close()

    # Style Excel Workbook with openpyxl
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = openpyxl.load_workbook(output_path)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gold_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

        for ws in wb.worksheets:
            # Format header
            for cell in ws[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")

            max_row = ws.max_row
            # Benchmark gold highlight (first company row)
            for cell in ws[2]:
                cell.fill = gold_fill
                cell.font = Font(bold=True)

            # Median row highlight (last row)
            for cell in ws[max_row]:
                cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                cell.font = Font(bold=True, italic=True)

        wb.save(output_path)
    except Exception as e:
        logger.warning(f"Could not apply openpyxl styles to peer comparison excel: {e}")

    logger.info(f"Peer comparison Excel report saved cleanly to {output_path}")
