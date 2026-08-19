"""
Valuation Module — Sprint 4 Day 26.

Computes:
- FCF yield (FCF / market_cap * 100) for all 92 latest-year companies
- Sector median P/E for each broad sector
- Overvaluation flags: Caution (P/E > sector_median * 1.5),
  Discount (P/E < sector_median * 0.7), Fair (otherwise)
- 5-year median P/E for each company

Generates:
- output/valuation_summary.xlsx (92 companies with all required columns)
- output/valuation_flags.csv (only Caution and Discount flagged companies)
"""

import os
import pandas as pd
import numpy as np
from src.utils.logger import get_logger
from src.db.connection import get_db_connection

logger = get_logger(__name__)


DB_PATH = "db/nifty100.db"
OUTPUT_VALUATION_XLSX = "output/valuation_summary.xlsx"
OUTPUT_FLAGS_CSV = "output/valuation_flags.csv"


def compute_valuation_summary(db_path: str = DB_PATH) -> pd.DataFrame:
    """
    Computes valuation metrics for all 92 latest-year companies.

    Returns DataFrame with columns:
    company_id, company_name, sector, P/E, P/B, EV/EBITDA,
    FCF_yield_pct, 5yr_median_PE, PE_vs_sector_median_pct, flag
    """
    logger.info("Starting Valuation Engine...")

    with get_db_connection(db_path) as conn:
        # ── Load latest-year universe ────────────────────────────────────────
        query_latest = """
            SELECT fr.company_id, fr.year, fr.pe_ratio, fr.pb_ratio,
                   fr.debt_to_equity, fr.free_cash_flow, fr.free_cash_flow_cr,
                   fr.interest_coverage,
                   c.company_name, c.ticker, s.sector_name,
                   pnl.sales, pnl.operating_profit, pnl.net_profit,
                   bs.total_assets, bs.total_equity, bs.total_liabilities,
                   mc.market_cap
            FROM financial_ratios fr
            JOIN companies c ON fr.company_id = c.company_id
            LEFT JOIN sectors s ON c.sector_id = s.sector_id
            LEFT JOIN profitandloss pnl ON fr.company_id = pnl.company_id AND fr.year = pnl.year
            LEFT JOIN balancesheet bs ON fr.company_id = bs.company_id AND fr.year = bs.year
            LEFT JOIN market_cap mc ON fr.company_id = mc.company_id AND CAST(mc.date AS INTEGER) = fr.year
            WHERE fr.year = (SELECT MAX(year) FROM financial_ratios fr2 WHERE fr2.company_id = fr.company_id)
        """
        latest_df = pd.read_sql_query(query_latest, conn)

        # ── Load 5-year P/E history for median calculation ───────────────────
        query_pe_hist = """
            SELECT fr.company_id, fr.year, fr.pe_ratio
            FROM financial_ratios fr
            WHERE fr.year >= (SELECT MAX(year) - 4 FROM financial_ratios)
        """
        pe_hist_df = pd.read_sql_query(query_pe_hist, conn)

    if latest_df.empty:
        logger.warning("No data found for valuation computation.")
        return pd.DataFrame()

    latest_df = latest_df.drop_duplicates(subset=["company_id"], keep="first")
    logger.info(f"Loaded {len(latest_df)} companies for valuation analysis.")

    # ── FCF Yield ────────────────────────────────────────────────────────────
    fcf_col = "free_cash_flow_cr" if "free_cash_flow_cr" in latest_df.columns else "free_cash_flow"
    latest_df["FCF_yield_pct"] = np.where(
        (latest_df["market_cap"].notna()) & (latest_df["market_cap"] > 0) & (latest_df[fcf_col].notna()),
        (latest_df[fcf_col] / latest_df["market_cap"]) * 100,
        np.nan
    )

    # ── EV/EBITDA Calculation ────────────────────────────────────────────────
    # EV = Market Cap + Total Debt - Cash (approximate using total_liabilities - total_equity as debt proxy)
    # EBITDA ≈ Operating Profit (approximation since raw EBITDA not available)
    latest_df["EV_EBITDA"] = np.where(
        (latest_df["market_cap"].notna()) & (latest_df["operating_profit"].notna()) &
        (latest_df["operating_profit"] > 0),
        latest_df["market_cap"] / latest_df["operating_profit"],
        np.nan
    )

    # ── 5-Year Median P/E ────────────────────────────────────────────────────
    median_pe_5yr = pe_hist_df.groupby("company_id")["pe_ratio"].median().reset_index()
    median_pe_5yr.columns = ["company_id", "5yr_median_PE"]
    latest_df = latest_df.merge(median_pe_5yr, on="company_id", how="left")

    # ── Sector Median P/E ────────────────────────────────────────────────────
    sector_median_pe = latest_df.groupby("sector_name")["pe_ratio"].median().reset_index()
    sector_median_pe.columns = ["sector_name", "sector_median_pe"]
    latest_df = latest_df.merge(sector_median_pe, on="sector_name", how="left")

    # P/E vs Sector Median %
    latest_df["PE_vs_sector_median_pct"] = np.where(
        (latest_df["pe_ratio"].notna()) & (latest_df["sector_median_pe"].notna()) &
        (latest_df["sector_median_pe"] > 0),
        ((latest_df["pe_ratio"] - latest_df["sector_median_pe"]) / latest_df["sector_median_pe"]) * 100,
        np.nan
    )

    # ── Valuation Flags ──────────────────────────────────────────────────────
    def assign_flag(row):
        pe = row.get("pe_ratio")
        sector_med = row.get("sector_median_pe")
        if pe is None or pd.isna(pe) or sector_med is None or pd.isna(sector_med) or sector_med <= 0:
            return "N/A"
        if pe > sector_med * 1.5:
            return "Caution"
        elif pe < sector_med * 0.7:
            return "Discount"
        else:
            return "Fair"

    latest_df["flag"] = latest_df.apply(assign_flag, axis=1)

    # ── Build Output DataFrame ───────────────────────────────────────────────
    output_df = latest_df[[
        "company_id", "company_name", "sector_name", "pe_ratio", "pb_ratio",
        "EV_EBITDA", "FCF_yield_pct", "5yr_median_PE",
        "PE_vs_sector_median_pct", "flag"
    ]].copy()

    output_df.columns = [
        "company_id", "company_name", "sector", "P/E", "P/B",
        "EV/EBITDA", "FCF_yield_pct", "5yr_median_PE",
        "PE_vs_sector_median_pct", "flag"
    ]

    # Round numeric columns
    for col in ["P/E", "P/B", "EV/EBITDA", "FCF_yield_pct", "5yr_median_PE", "PE_vs_sector_median_pct"]:
        if col in output_df.columns:
            output_df[col] = output_df[col].round(2)

    logger.info(f"Valuation summary computed for {len(output_df)} companies.")
    flag_counts = output_df["flag"].value_counts().to_dict()
    logger.info(f"Flag distribution: {flag_counts}")

    return output_df


def generate_valuation_outputs(output_df: pd.DataFrame = None):
    """Generates valuation_summary.xlsx and valuation_flags.csv."""
    if output_df is None:
        output_df = compute_valuation_summary()

    if output_df.empty:
        logger.warning("No valuation data to export.")
        return

    os.makedirs("output", exist_ok=True)

    # ── Excel Export ─────────────────────────────────────────────────────────
    output_df.to_excel(OUTPUT_VALUATION_XLSX, index=False, sheet_name="Valuation Summary")
    logger.info(f"Generated {OUTPUT_VALUATION_XLSX} with {len(output_df)} rows.")

    # ── Style the Excel with openpyxl ────────────────────────────────────────
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font

        wb = openpyxl.load_workbook(OUTPUT_VALUATION_XLSX)
        ws = wb.active

        # Header styling
        header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # Flag column color coding
        flag_col_idx = None
        for idx, cell in enumerate(ws[1], 1):
            if cell.value == "flag":
                flag_col_idx = idx
                break

        if flag_col_idx:
            caution_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            discount_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            fair_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")

            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                flag_cell = row[flag_col_idx - 1]
                if flag_cell.value == "Caution":
                    flag_cell.fill = caution_fill
                elif flag_cell.value == "Discount":
                    flag_cell.fill = discount_fill
                elif flag_cell.value == "Fair":
                    flag_cell.fill = fair_fill

        # Auto-width columns
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 30)

        wb.save(OUTPUT_VALUATION_XLSX)
    except Exception as e:
        logger.warning(f"Excel styling skipped: {e}")

    # ── CSV Export (Caution + Discount only) ─────────────────────────────────
    flags_df = output_df[output_df["flag"].isin(["Caution", "Discount"])].copy()
    flags_df.to_csv(OUTPUT_FLAGS_CSV, index=False)
    logger.info(f"Generated {OUTPUT_FLAGS_CSV} with {len(flags_df)} flagged companies.")


if __name__ == "__main__":
    val_df = compute_valuation_summary()
    generate_valuation_outputs(val_df)
    print(f"\n=== Valuation Complete ===")
    print(f"Total companies: {len(val_df)}")
    print(f"Flags: {val_df['flag'].value_counts().to_dict()}")
